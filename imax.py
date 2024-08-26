#! /usr/bin/python

"""
Reads and outputs the information from the USB port of the FrSky IMAX B6 mini charger.
Some notes are specific for Windows 10, 64 bit, running python 3.6.1, which was the development environment.
Whether this code will work on other operating systems is unknown; for sure the appropriate libusb.dll must be installed.
Requires:
    py_usb: wihch in turn requires libusb-1.0.dll be installed first. 
      The libusb-1.0 binary files can be downlaod as zip or 7z, but will not open in Windows directly because of security.
      Open 7-Zip and using the path window to the left, open up the 7z file. Migrate to the 64 directory 
      containing the already generated libusb-1.0 files.
      There are two subdirectries in the archive: for 64 bit or 32 bit windows systems. 
      Copy the three libusb.-1.0 files, including the .dll to the folder with this script.
      Copy these files to both C:/windows/system32 and C:/windows/syswow64 for 64 bit systems.
      Alternatively, this would have been the way to get at the dll and other files, if only using 
      the script folder, e.g., f:\rc_codes:
      import usb.backend.libusb1
      backend = usb.backend.libusb1.get_backend(find_library=lambda x: "F:\rc_codes\libusb-1.0.dll")
      dev = usb.core.find(backend=backend)
      Next install pysub using: pip install pyusb
      The tutorial on pyusb https://github.com/pyusb/pyusb/blob/master/docs/tutorial.rst

  openpyxl adds support to output directly and Excel .xslx file of the data.
  to get all usb devices run following from an independent python module
  There are many comments
"""

import os
import sys
import usb.core
import usb.util
import usb
import time
import datetime
import csv
import json
#from bok import text_update
from openpyxl import Workbook
from openpyxl.styles import Font, Fill


# Some settings for future use
battery_type = 'NIMH'
DC_mode = 'CHARGE'
cycle_num = 1
notes = 'write someting' #notes
settings_dict = {'max_chrg_time':"", 'max_input_V':"", 'cycle_delay':""}

# --------------DEVICE SPECIFIC SEARCH INFO-------------------------
# this substring is part of the USB/UART microcontroller product ID
my_device_usb = 'C8051F3xx' 
# -----------------------------------------------------------------

"""
#Generate the packets to output to the IMAX B6 Mini device. According to the Wireshack data,
#the packet size for in and out is 64 bytes to set up and collect data.
#Data out (host->imax) byte sequences below were obtained from WireShark USBPcap interrupt ENDPOINT_0UT
#The setup sequence consists of 11, 64 byte packets sent from the host to IMAX before querying for run data. 
#The first 10 INTERRUPT_OUT set up packets only change in the first 5 bytes of the 64 byte packets, 
#followed by an eleventh packet with many more byte changes.
#WireSharks' "Leftover Capture Data" sequences,with only a subset consisting of first 7 bytes.
out_setup = [[0x0f, 0x03, 0x5a, 0x00, 0x5a, 0xff, 0xff],
             [0x0f, 0x03, 0x57, 0x00, 0x57, 0xff, 0xff],
             [0x0f, 0x03, 0x57, 0x00, 0x57, 0xff, 0xff],
             [0x0f, 0x03, 0x5f, 0x00, 0x5f, 0xff, 0xff],
             [0x0f, 0x03, 0xfe, 0x00, 0xfe, 0xff, 0xff],
             [0x0f, 0x03, 0x5a, 0x00, 0x5a, 0xff, 0xff],
             [0x0f, 0x03, 0x5a, 0x00, 0x5a, 0xff, 0xff],
             [0x0f, 0x03, 0x5a, 0x00, 0x5a, 0xff, 0xff],
             [0x0f, 0x03, 0x5f, 0x00, 0x5f, 0xff, 0xff]]
"""
def output_csv(overview, data_out, final_dict):
  #used to output the data in csv form; not used in this version, but works
  cur_dir = os.getcwd()
  csv_file = 'testdata.csv'
  #Best way to get the current path of the module and file
  list_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), csv_file)
  # check if path and old event file exist; if not create it
  if os.path.isfile(list_file):
    #create the file and populate first time, we know data exists
    print('creating output')
    with open(list_file,"wb", encoding='utf-8')  as f:
      writer=csv.writer(f, delimiter=",", lineterminator="\r\n")
      key_list = list(overview.keys())
      writer.writerow(key_list)
      alist = [overview[k] for k in key_list]
      writer.writerow(alist)
      key_list = list(final_dict.keys())
      writer.writerow(key_list)
      alist = [final_dict[k] for k in key_list]
      writer.writerow(alist)
      writer.writerows(data_out)
  print('csv file done.')   
      
class find_class(object):
#used only by find_hids; example found online
  def __init__(self, class_):
    self._class = class_

  def __call__(self, device):
    # first, let's check the device
    if device.bDeviceClass == self._class:
      return True
    # ok, traverse all devices to find an
    # interface that matches our class
    try:
      for cfg in device:
        #find_descriptor: what's it?
        intf = usb.util.find_descriptor(cfg, bInterfaceClass=self._class)
        if intf is not None:
          return True
    except:
      pass
    return False

def find_hids():
  # find HID USB devices ONLY; standard usb code for Human Interface Devices (HID) is 0x03
  dev = usb.core.find(find_all=1, custom_match=find_class(3))
  #Finds devices, but some have no identifiable info, or restrict it
  return dev

def find_my_device():
  devicefnd = False
  my_devices = find_hids()
  if my_devices:
    msg = 'No Imax device found; check if connected, or not on.' 
    for d in my_devices:
      #print('d is: ', d)
      #attempt to get the product stuff
      try: 
        d._product = usb.util.get_string(d, d.iProduct)
      except Exception as e:
        print(f"Failed to get HID Product string: " + str(e))
      if d._product:
        print(d._product)
        #my_device_usb product variable is globally defined
        if my_device_usb in d._product:
          devicefnd = True
          msg = 'imax usb interface device: ' + d._product + ' found.'
          print('my devices found it')
          return msg
    #return default not found message, if not found
    print('find_my_devices says did not find it')
    return msg
  else: #no devices at all located
    return "No Devices found" 
 
def connect_imax(): 
  print('in connect_max')
  # This is over-engineered; to get here, we first check for device in find_my_device()
  # This is specific vendor and product ids for the FrSky IMAX B6 mini with USB port
  # Values from preceding methods or Wireshark usb output
  dev = usb.core.find(idVendor=0x0000, idProduct=0x0001)
  if dev:
    #print('device connnected: ', dev)
    print('Hexadecimal VendorID=' + hex(dev.idVendor) + ' & ProductID=' + hex(dev.idProduct) + '\n\n') 
    try:
      manuf = usb.util.get_string(dev, dev.iManufacturer)
      if manuf:
        print('manufacturer is: ', manuf)
    except:
      pass
    msg = manuf + ' Device connected: Hexadecimal VendorID=' + hex(dev.idVendor) + ' & ProductID=' + hex(dev.idProduct)

  try:
    # Linux support, detach kernel driver if used
    if dev.is_kernel_driver_active(0):
      try:
        dev.detach_kernel_driver(0)
        print("Kernel driver detached")
      except usb.USBError as e:
        raise IOError("Could not detach kernel driver") from e
    else:
      print("No kernel driver attached")
  except NotImplementedError as e:
    print(f"Not implemented: '{e}', proceeding")

  return dev


def start_imax(device_dict, settings_packet, always_reconnect):
  global settings_dict
  print('in start_max ')

  if not always_reconnect:
    dev = device_dict['device']
    initialized = False
    if dev is not None:
      try:
        dev.get_active_configuration()
        initialized = True
      except usb.USBError:
        print('Failed to communicate, probably disconnected')
        dev = None
  else:
    initialized = False
    dev = None

  if dev is None:
    dev = connect_imax()
    initialized = False
  if not dev:  # send back dummy tuple
    device_dict = {'device': None, 'EndPt_out': None, 'EndPt_in': None}
    read_data = {}
    settings_dict = {}
    data_out_packet = []
    return device_dict, read_data, settings_dict, data_out_packet
  if not initialized:
    # set the active configuration. With no arguments, the first configuration will be the active one
    print('Getting config')
    dev.set_configuration()
    print('Getting active configuration.')
    cfg = dev.get_active_configuration()
    print('Got active Cfg')
    print('Attempting to get default interface')
    intf = cfg[(0, 0)]
    print('Got interface')
    # AN ENDPOINT IS A BYTE BUCKET TO SEND OR RECEIVE DATA
    # match the first OUT endpoint
    # from the pyusb example
    # find the two endpoints; actually find endpoint attributes group; read and write take all the info under the ENDPOINT attributes group
    ep_out = usb.util.find_descriptor(intf, custom_match=lambda e: usb.util.endpoint_direction(
      e.bEndpointAddress) == usb.util.ENDPOINT_OUT)
    ep_in = usb.util.find_descriptor(intf, custom_match=lambda e: usb.util.endpoint_direction(
      e.bEndpointAddress) == usb.util.ENDPOINT_IN)
  else:
    ep_out = device_dict['EndPt_out']
    ep_in = device_dict['EndPt_in']

  assert ep_out is not None
  assert ep_in is not None
  #print('ep out is:', ep_out) #should be 0x1 + all Endpoint_out attributes
  #print('ep in is:', ep_in) #should be 0x81 + all other Endpoint_in attributes
  
  device_dict = {'device': dev, 'EndPt_out': ep_out, 'EndPt_in': ep_in}
  #print('start_imax:device_dict is: ', device_dict)
  imax_settings_out =[0x0f, 0x03, 0x5a, 0x00, 0x5a, 0xff, 0xff] + [0]*57 
  start_packet = [0x0f, 0x03, 0x5f, 0x00, 0x5f, 0xff, 0xff] + [0]*57  
  data_out_packet = [0x0f, 0x03, 0x55, 0x00, 0x55, 0xff, 0xff] + [0]*57
  stop_packet = [0x0f,	0x03, 0xfe,	0x000, 0xfe, 0xff, 0xff] + [0]*57  
  w_out = dev.write(ep_out, imax_settings_out)
  idle_settings = dev.read(ep_in.bEndpointAddress,ep_in.wMaxPacketSize) 
  #print('idle_settings are: ', idle_settings) 
  data_run = []
  if idle_settings:
    w_out = dev.write(ep_out, imax_settings_out)
    imax_sys_settings = dev.read(ep_in.bEndpointAddress,ep_in.wMaxPacketSize)
    cycle_delay = str(imax_sys_settings[5]) #in min
    max_chrg_time = str(imax_sys_settings[7]*256 + imax_sys_settings[8]) #in min
    max_mv = str(imax_sys_settings[10]*256 + imax_sys_settings[11]) #in mV
    max_input_V = str((imax_sys_settings[14]*256 + imax_sys_settings[15])/1000) #in V
    bat_V = (imax_sys_settings[18]*256 + imax_sys_settings[19])/1000
    print('Settings: Max Charge(mah): ' + max_chrg_time + ' min; mV(max): ' + max_input_V +' V; ' + 'Cycle delay: ' + cycle_delay + ' min.')
    print('Current battery voltage: ', bat_V)
  settings_dict = {'max_chrg_time':max_chrg_time, 'max_input_V':max_input_V, 'cycle_delay':cycle_delay, 'bat_V':bat_V}
  # Send what is some sort of start packet
  w_out = dev.write(ep_out, start_packet)
  # Write the run settings to the Imax
  if len(settings_packet) > 30:
    w_set = dev.write(ep_out,settings_packet)
    print('settings sent')
  else:
    print('no settings_packet transferred.')
    sys.exit()
  # for some reason Chargemaster sends 3 packets in quick succession; so us too....
  for i in range(3):
    w_out = dev.write(ep_out, data_out_packet)
  t = time.perf_counter()
  query_interval = 10
  #alternate input method if needed: - faster?
  #import array
  #in_buff = array.array('B', [0])
  #w_out = dev.write(ep_out, data_out_packet)
  print('sending first data packet')  
  w_out = dev.write(ep_out, data_out_packet) 
  try:
    data = dev.read(ep_in.bEndpointAddress,ep_in.wMaxPacketSize)  # using more general form of Endpoint_IN attributes
  except Exception as e: 
    print('Something went wrong: error is: ', e)
  else:
    run_status = []
    mah = []
    timer = []
    volts = []
    current = []
    ext_T = []
    internal_T = []
    cell1 = []
    cell2 = []
    cell3 = []
    cell4 = []
    cell5 = []
    cell6 = []
    read_data = {'run_status':run_status, 'mah':mah, 'timer': timer, 'volts':volts, 'current':current,
              'ext_T':ext_T, 'internal_T':internal_T, 'cell1':cell1, 'cell2':cell2,
              'cell3':cell3, 'cell4':cell4, 'cell5':cell5,  'cell6':cell6}

    read_data['run_status'].append(int(str(data[4])))
    read_data['mah'].append(int(str(data[5]*256 + data[6])))       #capacity, mah
    read_data['timer'].append(int(str(data[7]*256 + data[8])))     #seconds
    read_data['volts'].append(int(str(data[9]*256 + data[10])))    #volts
    read_data['current'].append(int(str(data[11]*256 + data[12]))) #amps
    read_data['ext_T'].append(int(str(data[13])))
    read_data['internal_T'].append(int(str(data[14])))             #deg. C?
    read_data['cell1'].append(int(str(data[17]*256 + data[18])))
    read_data['cell2'].append(int(str(data[19]*256 + data[20])))
    read_data['cell3'].append(int(str(data[21]*256 + data[22])))
    read_data['cell4'].append(int(str(data[23]*256 + data[24])))
    read_data['cell5'].append(int(str(data[25]*256 + data[26])))
    read_data['cell6'].append(int(str(data[27]*256 + data[28])))
 
    # print the data; this is same sequence that Milek7 used with hidapi; much appreciated effort.
    print ("state, energy, timer, voltage, current, ext temp, int temp, cell 1, cell 2, cell 3, cell 4, cell 5, cell 6")    
    print(
      str(data[4]) + ", " +                                 #state
      str(data[5] * 256 + data[6]) + ", " +                 #energy
      str(data[7] * 256 + data[8]) + ", " +                 #timer     
      str((data[9] * 256 + data[10]) / 1000.0) + ", " +     #voltage
      str((data[11] * 256 + data[12]) / 1000.0) + ", " +    #current
      str(data[13]) + ", " +                                #ext temp
      str(data[14]) + ", " +                                #int temp
      str((data[17] * 256 + data[18]) / 1000.0) + ", " +    #cell 1    
      str((data[19] * 256 + data[20]) / 1000.0) + ", " +    #cels 2
      str((data[21] * 256 + data[22]) / 1000.0) + ", " +    #cell 3
      str((data[23] * 256 + data[24]) / 1000.0) + ", " +    #cell 4
      str((data[25] * 256 + data[26]) / 1000.0) + ", " +    #cell 5
      str((data[27] * 256 + data[28]) / 1000.0)             #cell 6
    )
    #alternate conversion
    #print(int(str('0x'+data[5]+data[6])))
    data_run.append(data)
    sys.stdout.flush()
    return device_dict, read_data, settings_dict, data_out_packet

# Copied here just for easy Excel referencing:
  #run_modes = dict('bat_type':bat_type, 'chrg_type':chrg_type, 'nominal_mah':nominal_mah, 
  #                 'DC_or_CD':DC_or_CD, 'cycles':cycles, 'cells':cells, 'run_text':text_input}
  #final_out = {'final_mah':"", 'final_t':"", 'final_V':"", 'final_T':""}
  #data from read_data = {'mah':mah, 'timer':timer, 'volts';volts, 'current':current, 
  #                       'ext_T':ext_T, 'internal_T':internal_T, 'cell1':cell1, 'cell2':cell2,
  #                       'cell3':cell3, 'cell4':cell4, 'cell5':cell5, 'cell6':cell16}  

def write_excel_file(run_modes, final_out, data, settings_dict):
  wb = Workbook()
  ws = wb.active #gets the default single sheet created
  ws['A1'] = 'Date:'
  ws.merge_cells('A1:B1')
  ws['C1'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
  #dictionaries input are not ordered, so specify order out for value list
  words = ['battery type:','Number of cells:', 'charge type:', 'nomimal mah:', 'Cycle Type:', 'No. of cycles:', 'Informaiton']
  keylist = ['bat_type', 'cells', 'chrg_type', 'nominal_mah', 'DC_or_CD', 'cycles',  'run_text']
  datakeys = ['mah', 'timer', 'volts', 'current', 'ext_T', 'internal_T', 'cell1', 'cell2', 'cell3', 'cell4', 'cell5', 'cell6']  
  rmlist = [run_modes[key] for key in keylist] 
  print(rmlist)
  #put mode info
  for i in range(len(rmlist)): 
    ws['A'+str(i+1)] =  words[i]
    ws.merge_cells('A'+str(i+1) + ':B'+str(i+1))
    ws['C'+str(i+1)] = run_modes[keylist[i]]

  #add the final values
  words = ['Final capacity:', 'Final voltage:', 'Total time:'] 
  valuelist = [final_out['final_mah'], final_out['final_V'], final_out['final_t']]
  for i in range(len(words)):
    ws['E'+ str(i+1)] =  words[i]
    ws.merge_cells('E'+str(i+1) + ':F'+str(i+1))
    ws['G'+ str(i+1)] = valuelist[i]    

  words = ['Max. Input V:', 'Cycle Delay:', 'Max. Charge Time:'] 
  valuelist = [settings_dict['max_input_V'], settings_dict['cycle_delay'], settings_dict['max_chrg_time']]
  for i in range(len(words)):
    ws['I'+ str(i+1)] =  words[i]
    ws.merge_cells('I'+str(i+1) + ':J'+str(i+1))
    ws['K'+ str(i+1)] = valuelist[i]  
  print('past max')
  #add the data steam
  firstrow = 10
  headers = ['Capacity mah', 'Time s', 'Millivolts', 'Current', 'Ext.T', 'internal T', 'Cell 1 mV', 'Cell 2 mV', 'Cell 3 mV', 'Cell 4 mV', 'Cell 5 mV', 'Cell 6 mV']
  #add the headers
  print('l of h: ', len(headers))
  for i in range(len(headers)):
    ws.cell(row = 9, column = i+1).value = headers[i]
  #add the data
  print('trying data read')
  print('data is:', data)
  for c in range(len(datakeys)):
    alist = data[datakeys[c]]
    print(alist)
    for r in range(len(alist)):
      ws.cell(row = r + firstrow, column = c + 1).value = alist[r]

  wb.save('imax_run_' + datetime.datetime.now().strftime("%B_%d_%Y_%H_%M") + '.xlsx')
  return 'imax_run_' + datetime.datetime.now().strftime("%B_%d_%Y_%H_%M") + '.xlsx'
  
  #CSV data can be added; see imax_logger
